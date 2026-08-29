import logging

from fastapi import HTTPException, status
from sqlalchemy import desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import settings
from app.modules.auth.model import Teacher, TeacherSubject, User
from app.modules.organization_structure.model import Kafedra
from app.modules.file.storage import public_url, store_upload
from app.modules.quiz.model import Question, Subject

from .schemas import (
    QuestionBulkDeleteRequest,
    QuestionCatalogResponse,
    QuestionCreateRequest,
    QuestionListRequest,
    QuestionListResponse,
    QuestionSubjectSummary,
    QuestionTeacherSummary,
)

logger = logging.getLogger(__name__)


class QuestionRepository:
    async def _assigned_subject_ids(self, session: AsyncSession, user: User) -> list[int]:
        """Oʻqituvchi biriktirilgan fanlar roʻyxati.

        Savollar bazasi shu boʻyicha koʻrsatiladi: oʻqituvchi oʻzi dars
        beradigan fanning savollarini koʻrishi kerak, garchi ularni hamkasbi
        yozgan boʻlsa ham. Tahrirlash esa faqat oʻz savolida qoladi.
        """
        rows = await session.scalars(
            select(TeacherSubject.subject_id)
            .join(Teacher, Teacher.id == TeacherSubject.teacher_id)
            .where(Teacher.user_id == user.id)
        )
        return list(rows)

    def _visible_questions_filter(self, user: User, subject_ids: list[int]):
        """Oʻqituvchiga koʻrinadigan savollar sharti.

        Oʻz savollari HAM kiradi: biriktirma oʻzgarsa (fan boshqaga oʻtsa)
        oʻqituvchi oʻzi yozgan savollarni yoʻqotib qoʻymasligi kerak.
        """
        own = Question.user_id == user.id
        if not subject_ids:
            return own
        return or_(own, Question.subject_id.in_(subject_ids))

    async def get_catalog(
        self, session: AsyncSession, current_user: User, search: str | None = None
    ) -> QuestionCatalogResponse:
        """Return teacher -> subject counts for the question-bank catalogue."""
        stmt = (
            select(
                Question.user_id.label("teacher_user_id"),
                User.username,
                Teacher.full_name,
                Kafedra.id.label("kafedra_id"),
                Kafedra.name.label("kafedra_name"),
                Subject.id.label("subject_id"),
                Subject.name.label("subject_name"),
                func.count(Question.id).label("question_count"),
            )
            .join(User, User.id == Question.user_id)
            .outerjoin(Teacher, Teacher.user_id == User.id)
            .outerjoin(Kafedra, Kafedra.id == Teacher.kafedra_id)
            .join(Subject, Subject.id == Question.subject_id)
            .where(Question.is_latest.is_(True), Question.is_active.is_(True))
        )
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin:
            subject_ids = await self._assigned_subject_ids(session, current_user)
            stmt = stmt.where(self._visible_questions_filter(current_user, subject_ids))
        if search:
            pattern = f"%{search}%"
            stmt = stmt.where(
                or_(
                    Teacher.full_name.ilike(pattern),
                    User.username.ilike(pattern),
                    Subject.name.ilike(pattern),
                )
            )
        stmt = stmt.group_by(
            Question.user_id,
            User.username,
            Teacher.full_name,
            Kafedra.id,
            Kafedra.name,
            Subject.id,
            Subject.name,
        ).order_by(Teacher.full_name.asc().nullslast(), User.username.asc(), Subject.name.asc())

        teachers: dict[int, QuestionTeacherSummary] = {}
        for row in (await session.execute(stmt)).all():
            teacher = teachers.get(row.teacher_user_id)
            if teacher is None:
                teacher = QuestionTeacherSummary(
                    teacher_user_id=row.teacher_user_id,
                    username=row.username,
                    full_name=row.full_name,
                    kafedra_id=row.kafedra_id,
                    kafedra_name=row.kafedra_name,
                    question_count=0,
                    subjects=[],
                )
                teachers[row.teacher_user_id] = teacher
            teacher.question_count += row.question_count
            teacher.subjects.append(
                QuestionSubjectSummary(
                    subject_id=row.subject_id,
                    subject_name=row.subject_name,
                    question_count=row.question_count,
                )
            )
        return QuestionCatalogResponse(teachers=list(teachers.values()))

    async def create_question(
        self, session: AsyncSession, data: QuestionCreateRequest, current_user: User
    ) -> Question:
        """Savol yaratadi. Muallif — har doim soʻrov yuborgan foydalanuvchi.

        Ilgari muallif soʻrov tanasidagi ``user_id`` dan olinardi va hech
        tekshirilmasdi: bir oʻqituvchi savolni boshqasining nomidan yozib
        qoʻyishi mumkin edi. Admin uchun esa bu imkoniyat qoladi — u savolni
        kerakli oʻqituvchiga biriktira oladi.
        """
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        author_id = data.user_id if is_admin and data.user_id else current_user.id

        if not is_admin:
            # Oʻz fanidan tashqariga savol qoʻshib boʻlmaydi. Biriktirmasi
            # umuman yoʻq oʻqituvchini bloklamaymiz — EduPlan sinxronizatsiyasi
            # kechikkan boʻlishi mumkin, bu esa ishlashni butunlay toʻxtatardi.
            subject_ids = await self._assigned_subject_ids(session, current_user)
            if subject_ids and data.subject_id not in subject_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Bu fan sizga biriktirilmagan — unga savol qoʻsholmaysiz",
                )

        new_question = Question(
            subject_id=data.subject_id,
            user_id=author_id,
            text=data.text,
            option_a=data.option_a,
            option_b=data.option_b,
            option_c=data.option_c,
            option_d=data.option_d,
            correct_option=data.correct_option,
            question_type=data.question_type.value,
            payload=data.payload,
        )
        session.add(new_question)

        try:
            await session.commit()
            await session.refresh(new_question)
        except Exception:
            await session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database error",
            )
        return new_question

    async def get_question(self, session: AsyncSession, question_id: int, current_user: User) -> Question:
        stmt = (
            select(Question)
            .options(
                selectinload(Question.subject),
                selectinload(Question.user),
            )
            .where(Question.id == question_id)
        )
        result = await session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")

        # Koʻrish — oʻz savoli yoki oʻzi dars beradigan fanning savoli.
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin and question.user_id != current_user.id:
            subject_ids = await self._assigned_subject_ids(session, current_user)
            if question.subject_id not in subject_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Bu savol sizning fanlaringizga tegishli emas",
                )

        return question

    async def list_questions(
        self, session: AsyncSession, request: QuestionListRequest, current_user: User
    ) -> QuestionListResponse:
        stmt = (
            select(Question)
            .options(
                selectinload(Question.subject),
                selectinload(Question.user),
            )
            .where(Question.is_latest.is_(True), Question.is_active.is_(True))
        )

        # Check if user is teacher (not admin)
        is_teacher = any(role.name.lower() == "teacher" for role in current_user.roles)
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)

        subject_ids: list[int] = []
        if not is_admin and is_teacher:
            # Oʻqituvchi biriktirilgan fanlarining savollarini koʻradi —
            # ilgari faqat oʻzi yozganini koʻrardi va oʻz fanining bazasi
            # unga boʻsh koʻrinardi.
            subject_ids = await self._assigned_subject_ids(session, current_user)
            stmt = stmt.where(self._visible_questions_filter(current_user, subject_ids))

        if request.text:
            stmt = stmt.where(Question.text.ilike(f"%{request.text}%"))

        if request.subject_id:
            stmt = stmt.where(Question.subject_id == request.subject_id)

        if request.user_id:
            stmt = stmt.where(Question.user_id == request.user_id)

        stmt = stmt.order_by(desc(Question.created_at))
        stmt = stmt.offset(request.offset).limit(request.limit)

        result = await session.execute(stmt)
        questions = result.scalars().all()

        count_stmt = (
            select(func.count()).select_from(Question).where(Question.is_latest.is_(True), Question.is_active.is_(True))
        )
        if not is_admin and is_teacher:
            count_stmt = count_stmt.where(self._visible_questions_filter(current_user, subject_ids))
        if request.text:
            count_stmt = count_stmt.where(Question.text.ilike(f"%{request.text}%"))
        if request.subject_id:
            count_stmt = count_stmt.where(Question.subject_id == request.subject_id)
        if request.user_id:
            count_stmt = count_stmt.where(Question.user_id == request.user_id)

        total_result = await session.execute(count_stmt)
        total = total_result.scalar() or 0

        return QuestionListResponse(total=total, page=request.page, limit=request.limit, questions=questions)

    async def update_question(
        self,
        session: AsyncSession,
        question_id: int,
        data: QuestionCreateRequest,
        current_user: User,
    ) -> Question:
        """Editing a question never mutates the row in place — it creates a new
        version, flips is_latest on the old one, and repoints any quiz_questions
        that referenced the old version onto the new one. Already-started attempts
        (UserAnswers reserved at start_quiz) keep pointing at the exact version the
        student was shown, so editing mid-flight never corrupts a live attempt."""
        stmt = (
            select(Question)
            .options(
                selectinload(Question.subject),
                selectinload(Question.user),
            )
            .where(Question.id == question_id)
        )
        result = await session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin and question.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: you can only update your own questions",
            )

        new_question = Question(
            subject_id=data.subject_id,
            user_id=data.user_id,
            text=data.text,
            option_a=data.option_a,
            option_b=data.option_b,
            option_c=data.option_c,
            option_d=data.option_d,
            correct_option=data.correct_option,
            question_type=data.question_type.value,
            payload=data.payload,
            original_question_id=question.original_question_id or question.id,
            version=question.version + 1,
            is_latest=True,
            is_active=question.is_active,
        )
        session.add(new_question)
        question.is_latest = False

        try:
            await session.flush()

            from app.modules.quiz.model import QuizQuestion

            await session.execute(
                QuizQuestion.__table__.update()
                .where(QuizQuestion.question_id == question.id)
                .values(question_id=new_question.id)
            )

            await session.commit()
            await session.refresh(new_question)
        except Exception:
            await session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database error",
            )

        return new_question

    async def delete_question(self, session: AsyncSession, question_id: int, current_user: User) -> None:
        stmt = select(Question).where(Question.id == question_id)
        result = await session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Question not found")

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin and question.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access denied: you can only delete your own questions",
            )

        # Soft delete: the row stays (it may already be referenced by
        # quiz_questions/user_answers) — it's just excluded from future selection.
        question.is_active = False
        await session.commit()

    async def bulk_delete_questions(
        self, session: AsyncSession, data: QuestionBulkDeleteRequest, current_user: User
    ) -> int:
        from sqlalchemy import update

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin:
            if data.user_id != current_user.id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Access denied: you can only delete your own questions",
                )

        stmt = (
            update(Question)
            .where(Question.subject_id == data.subject_id, Question.user_id == data.user_id)
            .values(is_active=False)
        )

        result = await session.execute(stmt)
        await session.commit()

        return result.rowcount

    async def upload_image(self, session: AsyncSession, file, current_user) -> str:
        """Savol rasmini yuklaydi va fayl kutubxonasiga qayd etadi.

        Papka avvalgidek `question/` — bazadagi mavjud havolalar shu yerga
        ishora qiladi va ularni buzib boʻlmaydi."""
        stored = await store_upload(
            session,
            file,
            owner_user_id=current_user.id if current_user else None,
            subdir="question",
        )
        await session.commit()
        await session.refresh(stored, ["blob"])
        return public_url(stored.blob.stored_path)

    async def upload_questions_excel(
        self, session: AsyncSession, file, subject_id: int, user_id: int
    ) -> list[Question]:
        import io

        import pandas as pd

        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # Verify there are at least 5 columns
        if len(df.columns) < 5:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain at least 5 columns (question, option A, option B, option C, option D)",
            )

        questions = []
        warnings = []
        for index, row in df.iterrows():
            # If subject_id is in row, use it, else use param
            # If image is in row, use it

            # Using positional indices instead of column names
            text = str(row.iloc[0]) if not pd.isna(row.iloc[0]) else ""
            opt_a = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ""
            opt_b = str(row.iloc[2]) if not pd.isna(row.iloc[2]) else ""
            opt_c = str(row.iloc[3]) if not pd.isna(row.iloc[3]) else ""
            opt_d = str(row.iloc[4]) if not pd.isna(row.iloc[4]) else ""

            q_subject_id = subject_id
            if "subject_id" in df.columns and not pd.isna(row["subject_id"]):
                try:
                    q_subject_id = int(row["subject_id"])
                except (ValueError, TypeError):
                    pass

            correct_option = "a"
            if "correct_option" in df.columns and not pd.isna(row["correct_option"]):
                candidate = str(row["correct_option"]).strip().lower()
                if candidate in ("a", "b", "c", "d"):
                    correct_option = candidate
                else:
                    warnings.append(f"Qator {index + 2}: to'g'ri javob noto'g'ri ko'rsatilgan, 'A' ishlatildi")
            else:
                warnings.append(f"Qator {index + 2}: to'g'ri javob ko'rsatilmagan, 'A' ishlatildi")

            question = Question(
                subject_id=q_subject_id,
                user_id=user_id,
                text=text,
                option_a=opt_a,
                option_b=opt_b,
                option_c=opt_c,
                option_d=opt_d,
                correct_option=correct_option,
            )
            questions.append(question)

        session.add_all(questions)

        try:
            await session.commit()
        except Exception:
            await session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database error during bulk upload",
            )

        return {"questions": questions, "warnings": warnings}

    async def download_questions_excel(
        self,
        session: AsyncSession,
        subject_id: int | None = None,
        user_id: int | None = None,
        text: str | None = None,
    ) -> bytes:
        import io
        import re

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        def strip_html(html: str) -> str:
            """Remove HTML tags and return plain text."""
            clean = re.sub(r"<[^>]+>", "", html or "")
            return clean.strip()

        # Query all matching questions (no pagination)
        stmt = (
            select(Question)
            .options(
                selectinload(Question.subject),
                selectinload(Question.user),
            )
            .where(Question.is_latest.is_(True), Question.is_active.is_(True))
        )

        if text:
            stmt = stmt.where(Question.text.ilike(f"%{text}%"))
        if subject_id:
            stmt = stmt.where(Question.subject_id == subject_id)
        if user_id:
            stmt = stmt.where(Question.user_id == user_id)

        stmt = stmt.order_by(desc(Question.created_at))

        result = await session.execute(stmt)
        questions = result.scalars().all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Savollar"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "№",
            "Savol",
            "A variant",
            "B variant",
            "C variant",
            "D variant",
            "To'g'ri javob",
            "Fan",
            "Foydalanuvchi",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx, q in enumerate(questions, 2):
            subject_name = q.subject.name if q.subject else "-"
            username = q.user.username if q.user else "-"

            values = [
                row_idx - 1,
                strip_html(q.text),
                strip_html(q.option_a),
                strip_html(q.option_b),
                strip_html(q.option_c),
                strip_html(q.option_d),
                q.correct_option.upper(),
                subject_name,
                username,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 6  # №
        ws.column_dimensions["B"].width = 50  # Savol
        ws.column_dimensions["C"].width = 25  # A
        ws.column_dimensions["D"].width = 25  # B
        ws.column_dimensions["E"].width = 25  # C
        ws.column_dimensions["F"].width = 25  # D
        ws.column_dimensions["G"].width = 15  # To'g'ri javob
        ws.column_dimensions["H"].width = 20  # Fan
        ws.column_dimensions["I"].width = 18  # Foydalanuvchi

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


get_question_repository = QuestionRepository()
